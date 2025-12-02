#%%By Janis 250226
"""
Excel文件处理工具模块（函数式）
功能：
- 读取CSV/XLSX文件转为DataFrame
- 将cross-section数据转换为数据库三列表格式
- 批量文件操作
- 文件夹分类和目录树生成
- Excel错误检查（错行、空值、异常值）
"""

import pandas as pd
import numpy as np
import os
import logging
from pathlib import Path
from typing import Dict, List, Union, Tuple, Optional
from datetime import datetime
import json
from .config import get_excel_config


def _get_default_logger():
    """获取默认logger"""
    logger = logging.getLogger('ExcelProcessor')
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        ch.setLevel(logging.WARNING)
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        ch.setFormatter(formatter)
        logger.addHandler(ch)
    return logger


def read_csv_with_encoding(filepath: Path, logger: Optional[logging.Logger] = None, **kwargs) -> pd.DataFrame:
    """
    尝试使用多种编码读取CSV文件
    
    Args:
        filepath: CSV文件路径
        logger: 日志记录器，如果为None则使用默认logger
        **kwargs: 传递给pd.read_csv的额外参数
        
    Returns:
        DataFrame
        
    Raises:
        Exception: 所有编码尝试失败
    """
    if logger is None:
        logger = _get_default_logger()
    
    # 如果用户已指定encoding，直接使用
    if 'encoding' in kwargs:
        return pd.read_csv(filepath, **kwargs)
    
    # 从配置文件获取编码列表
    excel_config = get_excel_config()
    encodings = excel_config.get('encodings', [
        'utf-8', 'utf-8-sig', 'gb18030', 'gbk', 'gb2312', 'latin1', 'cp936', 'big5'
    ])
    
    last_error = None
    
    for encoding in encodings:
        try:
            df = pd.read_csv(filepath, encoding=encoding, **kwargs)
            logger.info(f"使用编码 '{encoding}' 成功读取CSV文件: {filepath}")
            return df
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            logger.debug(f"编码 '{encoding}' 读取失败，尝试下一个编码")
            continue
        except Exception as e:
            # 其他错误（如文件格式错误）直接抛出
            logger.error(f"读取CSV文件时发生错误: {str(e)}")
            raise
    
    # 所有编码都失败
    error_msg = f"无法读取文件 {filepath}，已尝试编码: {', '.join(encodings)}"
    logger.error(error_msg)
    if last_error:
        raise Exception(f"{error_msg}. 最后一个错误: {str(last_error)}")
    else:
        raise Exception(error_msg)


def read_file(
    filepath: Union[str, Path], 
    logger: Optional[logging.Logger] = None,
    **kwargs
) -> pd.DataFrame:
    """
    读取CSV或XLSX文件为DataFrame
    支持多种编码格式（UTF-8, GB2312, GBK, GB18030等）
    
    Args:
        filepath: 文件路径
        logger: 日志记录器，如果为None则使用默认logger
        **kwargs: 传递给pd.read_csv或pd.read_excel的额外参数
        
    Returns:
        DataFrame
        
    Raises:
        ValueError: 不支持的文件格式
        FileNotFoundError: 文件不存在
    """
    if logger is None:
        logger = _get_default_logger()
    
    filepath = Path(filepath)
    
    if not filepath.exists():
        logger.error(f"文件不存在: {filepath}")
        raise FileNotFoundError(f"文件不存在: {filepath}")
    
    try:
        if filepath.suffix.lower() in ['.csv']:
            # CSV文件需要处理编码问题
            df = read_csv_with_encoding(filepath, logger=logger, **kwargs)
            logger.info(f"成功读取CSV文件: {filepath}, 形状: {df.shape}")
        elif filepath.suffix.lower() in ['.xlsx', '.xls']:
            # 尝试读取Excel文件，处理openpyxl样式问题
            try:
                df = pd.read_excel(filepath, engine='openpyxl', **kwargs)
                logger.info(f"成功读取Excel文件: {filepath}, 形状: {df.shape}")
            except (TypeError, ValueError, Exception) as style_error:
                # 如果遇到样式错误（如命名样式为None），尝试只读取数据
                if 'NamedCellStyle' in str(style_error) or 'name' in str(style_error):
                    logger.warning(f"Excel文件存在样式问题，使用备选方案读取: {filepath}")
                    
                    # 方案1: 尝试使用 xlrd 引擎（对于 .xls 文件）
                    if filepath.suffix.lower() == '.xls':
                        try:
                            df = pd.read_excel(filepath, engine='xlrd', **kwargs)
                            logger.info(f"使用xlrd引擎成功读取: {filepath}, 形状: {df.shape}")
                        except Exception as xlrd_error:
                            logger.error(f"xlrd引擎失败: {xlrd_error}")
                            raise style_error
                    else:
                        # 方案2: 使用 pyxlsb 或其他库
                        try:
                            # 尝试使用 calamine (python-calamine) - 更快且不处理样式
                            try:
                                import python_calamine
                                wb = python_calamine.CalamineWorkbook.from_path(str(filepath))
                                ws = wb.get_sheet_by_index(0)
                                data = ws.to_python()
                                if data:
                                    cols = data[0]
                                    df = pd.DataFrame(data[1:], columns=cols)
                                else:
                                    df = pd.DataFrame()
                                logger.info(f"使用calamine成功读取: {filepath}, 形状: {df.shape}")
                            except (ImportError, Exception):
                                # 方案3: 手动使用openpyxl但捕获样式错误
                                import openpyxl
                                from openpyxl import load_workbook
                                
                                # 猴子补丁：临时修复 NamedCellStyle
                                original_init = openpyxl.styles.named_styles._NamedCellStyle.__init__
                                def patched_init(self, name=None, **kw):
                                    if name is None:
                                        name = "Normal"  # 使用默认名称
                                    original_init(self, name=name, **kw)
                                
                                openpyxl.styles.named_styles._NamedCellStyle.__init__ = patched_init
                                
                                try:
                                    wb = load_workbook(filepath, data_only=True, read_only=False)
                                    ws = wb.active
                                    data = [[cell.value for cell in row] for row in ws.iter_rows()]
                                    if data:
                                        cols = data[0]
                                        df = pd.DataFrame(data[1:], columns=cols)
                                    else:
                                        df = pd.DataFrame()
                                    wb.close()
                                    logger.info(f"使用openpyxl(patched)成功读取: {filepath}, 形状: {df.shape}")
                                finally:
                                    # 恢复原始方法
                                    openpyxl.styles.named_styles._NamedCellStyle.__init__ = original_init
                        except Exception as backup_error:
                            logger.error(f"所有备选方案失败: {backup_error}")
                            raise style_error
                else:
                    raise
        else:
            raise ValueError(f"不支持的文件格式: {filepath.suffix}")
        
        return df
        
    except Exception as e:
        logger.error(f"读取文件失败 {filepath}: {str(e)}")
        raise


def cross_section_to_db_format(
    df: pd.DataFrame, 
    key_columns: List[str],
    value_columns: List[str],
    key_value_mapping: Dict[str, str],
    additional_fields: Optional[Dict[str, any]] = None,
    logger: Optional[logging.Logger] = None
) -> pd.DataFrame:
    """
    将cross-section格式数据转换为数据库三列表格式
    
    Args:
        df: 输入DataFrame
        key_columns: 键列（如code, name等，保持不变的列）
        value_columns: 值列（需要转换的列，如各个日期或指标）
        key_value_mapping: 列名映射，如{'variable': 'metric', 'value': 'value'}
        additional_fields: 额外添加的字段，如{'datetime': '2024-01-01'}
        logger: 日志记录器，如果为None则使用默认logger
            
    Returns:
        转换后的DataFrame，格式为：[key_columns, metric, value, ...]
        
    Example:
        输入:
            code | name | 2024-01-01 | 2024-01-02
            A001 | 股票A |    100     |    101
            
        输出:
            code | name | date       | value
            A001 | 股票A | 2024-01-01 |  100
            A001 | 股票A | 2024-01-02 |  101
    """
    if logger is None:
        logger = _get_default_logger()
    
    try:
        # 使用pd.melt进行转换
        df_melted = pd.melt(
            df,
            id_vars=key_columns,
            value_vars=value_columns,
            var_name='variable',
            value_name='value'
        )
        
        # 应用列名映射
        if key_value_mapping:
            df_melted.rename(columns=key_value_mapping, inplace=True)
        
        # 添加额外字段
        if additional_fields:
            for field_name, field_value in additional_fields.items():
                df_melted[field_name] = field_value
        
        logger.info(f"Cross-section转换完成: {df.shape} -> {df_melted.shape}")
        return df_melted
        
    except Exception as e:
        logger.error(f"Cross-section转换失败: {str(e)}")
        raise


def batch_read_files(
    folder_path: Union[str, Path],
    file_pattern: str = "*.csv",
    recursive: bool = False,
    logger: Optional[logging.Logger] = None,
    **read_kwargs
) -> Dict[str, pd.DataFrame]:
    """
    批量读取文件夹中的文件
    
    Args:
        folder_path: 文件夹路径
        file_pattern: 文件匹配模式，如"*.csv", "*.xlsx"
        recursive: 是否递归搜索子文件夹
        logger: 日志记录器，如果为None则使用默认logger
        **read_kwargs: 传递给read_file的参数
        
    Returns:
        字典，键为文件路径，值为DataFrame
    """
    if logger is None:
        logger = _get_default_logger()
    
    folder_path = Path(folder_path)
    
    if not folder_path.exists():
        logger.error(f"文件夹不存在: {folder_path}")
        raise FileNotFoundError(f"文件夹不存在: {folder_path}")
    
    # 查找文件
    if recursive:
        files = list(folder_path.rglob(file_pattern))
    else:
        files = list(folder_path.glob(file_pattern))
    
    logger.info(f"找到 {len(files)} 个文件匹配 '{file_pattern}'")
    
    # 读取文件
    results = {}
    errors = {}
    
    for file in files:
        try:
            df = read_file(file, logger=logger, **read_kwargs)
            results[str(file)] = df
        except Exception as e:
            logger.warning(f"读取文件失败 {file}: {str(e)}")
            errors[str(file)] = str(e)
    
    logger.info(f"成功读取 {len(results)}/{len(files)} 个文件")
    
    if errors:
        logger.warning(f"失败文件列表: {list(errors.keys())}")
    
    return results


def batch_write_files(
    data_dict: Dict[str, pd.DataFrame],
    output_dir: Union[str, Path],
    file_format: str = 'csv',
    create_subdirs: bool = True,
    logger: Optional[logging.Logger] = None,
    **write_kwargs
) -> List[str]:
    """
    批量写入文件
    CSV默认使用utf-8-sig编码，确保Excel能正确显示中文
    
    Args:
        data_dict: 字典，键为相对路径/文件名，值为DataFrame
        output_dir: 输出根目录
        file_format: 输出格式，'csv'或'xlsx'
        create_subdirs: 是否创建子目录
        logger: 日志记录器，如果为None则使用默认logger
        **write_kwargs: 传递给to_csv或to_excel的参数
        
    Returns:
        成功写入的文件路径列表
    """
    if logger is None:
        logger = _get_default_logger()
    
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    success_files = []
    
    for relative_path, df in data_dict.items():
        try:
            # 构建输出路径
            output_path = output_dir / relative_path
            
            # 确保后缀正确
            if file_format == 'csv' and not str(output_path).endswith('.csv'):
                output_path = output_path.with_suffix('.csv')
            elif file_format == 'xlsx' and not str(output_path).endswith('.xlsx'):
                output_path = output_path.with_suffix('.xlsx')
            
            # 创建子目录
            if create_subdirs:
                output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 写入文件
            if file_format == 'csv':
                # 默认使用utf-8-sig编码（带BOM），Excel能正确识别
                if 'encoding' not in write_kwargs:
                    write_kwargs['encoding'] = 'utf-8-sig'
                df.to_csv(output_path, **write_kwargs)
            elif file_format == 'xlsx':
                df.to_excel(output_path, **write_kwargs)
            else:
                raise ValueError(f"不支持的文件格式: {file_format}")
            
            success_files.append(str(output_path))
            logger.info(f"成功写入文件: {output_path}")
            
        except Exception as e:
            logger.error(f"写入文件失败 {relative_path}: {str(e)}")
    
    logger.info(f"成功写入 {len(success_files)}/{len(data_dict)} 个文件")
    return success_files


def create_directory_tree(
    data_dict: Dict[str, pd.DataFrame],
    output_dir: Union[str, Path],
    categorize_by: Optional[str] = None,
    logger: Optional[logging.Logger] = None
) -> Dict[str, List[str]]:
    """
    创建目录树并分类保存文件
    
    Args:
        data_dict: 数据字典
        output_dir: 输出目录
        categorize_by: 分类依据，如'date', 'metric'等（DataFrame中的列名）
        logger: 日志记录器，如果为None则使用默认logger
        
    Returns:
        目录树字典，键为类别，值为文件路径列表
    """
    if logger is None:
        logger = _get_default_logger()
    
    output_dir = Path(output_dir)
    directory_tree = {}
    
    if categorize_by is None:
        # 不分类，直接保存
        for name, df in data_dict.items():
            category = 'all'
            if category not in directory_tree:
                directory_tree[category] = []
            directory_tree[category].append(name)
    else:
        # 按指定列分类
        for name, df in data_dict.items():
            if categorize_by in df.columns:
                categories = df[categorize_by].unique()
                for category in categories:
                    if pd.isna(category):
                        category = 'unknown'
                    else:
                        category = str(category)
                    
                    if category not in directory_tree:
                        directory_tree[category] = []
                    
                    # 创建子目录
                    subdir = output_dir / category
                    subdir.mkdir(parents=True, exist_ok=True)
                    
                    # 筛选数据
                    df_subset = df[df[categorize_by] == category]
                    
                    # 保存文件
                    file_path = subdir / f"{Path(name).stem}_{category}.csv"
                    df_subset.to_csv(file_path, index=False, encoding='utf_8_sig')
                    
                    directory_tree[category].append(str(file_path))
                    logger.info(f"保存文件到分类 '{category}': {file_path}")
    
    # 保存目录树结构
    tree_file = output_dir / "directory_tree.json"
    with open(tree_file, 'w', encoding='utf-8') as f:
        json.dump(directory_tree, f, ensure_ascii=False, indent=2)
    
    logger.info(f"目录树已保存到: {tree_file}")
    return directory_tree


def apply_time_alignment(
    df: pd.DataFrame,
    date_column: str = '日期',
    metric_column: str = 'variable',
    open_metric_names: Optional[set] = None,
    open_time: Optional[str] = None,
    other_time: Optional[str] = None,
    inplace: bool = False,
    logger: Optional[logging.Logger] = None
) -> pd.DataFrame:
    """
    根据指标类型为日期列添加时间戳
    
    开盘价在交易日09:30可用，其他价格/成交量在15:00收盘后可用。
    
    Args:
        df: DataFrame
        date_column: 日期列名
        metric_column: 指标列名（用于判断是否为开盘价）
        open_metric_names: 开盘价指标名称集合，默认从config.json读取
        open_time: 开盘价对应的时间，默认从config.json读取
        other_time: 其他指标对应的时间，默认从config.json读取
        inplace: 是否原地修改
        logger: 日志记录器，如果为None则使用默认logger
        
    Returns:
        添加时间戳后的DataFrame
        
    Example:
        >>> df = pd.DataFrame({
        ...     '日期': ['2024-01-01', '2024-01-01'],
        ...     'variable': ['开盘价(元)', '收盘价(元)'],
        ...     'value': [100, 101]
        ... })
        >>> df = apply_time_alignment(df)
        >>> df['日期']
        0    2024-01-01 09:30:01
        1    2024-01-01 15:00:01
    """
    if logger is None:
        logger = _get_default_logger()
    
    if not inplace:
        df = df.copy()
    
    # 从配置文件获取时间对齐参数
    excel_config = get_excel_config()
    time_alignment_config = excel_config.get('time_alignment', {})
    
    if open_time is None:
        open_time = time_alignment_config.get('open_time', '09:30:01')
    if other_time is None:
        other_time = time_alignment_config.get('other_time', '15:00:01')
    if open_metric_names is None:
        open_metric_names = set(time_alignment_config.get('open_metric_names', ['开盘价', '开盘价(元)']))
    
    if date_column not in df.columns:
        logger.error(f"日期列 '{date_column}' 不存在")
        return df
    
    if metric_column not in df.columns:
        logger.warning(f"指标列 '{metric_column}' 不存在，将所有行使用 {other_time}")
        df[date_column] = df[date_column].astype(str) + " " + other_time
        df[date_column] = pd.to_datetime(df[date_column])
        return df
    
    try:
        # 创建时间戳
        open_mask = df[metric_column].isin(open_metric_names)
        
        # 为开盘价添加09:30时间戳
        df.loc[open_mask, date_column] = df.loc[open_mask, date_column].astype(str) + " " + open_time
        
        # 为其他指标添加15:00时间戳
        df.loc[~open_mask, date_column] = df.loc[~open_mask, date_column].astype(str) + " " + other_time
        
        # 转换为datetime类型
        df[date_column] = pd.to_datetime(df[date_column])
        
        logger.info(
            f"时间对齐完成: {open_mask.sum()} 行 → {open_time}, "
            f"{(~open_mask).sum()} 行 → {other_time}"
        )
        
    except Exception as e:
        logger.error(f"时间对齐失败: {str(e)}")
        raise
    
    return df


def check_excel_errors(
    df: pd.DataFrame,
    checks: Optional[Dict[str, any]] = None,
    logger: Optional[logging.Logger] = None
) -> Tuple[bool, List[Dict]]:
    """
    检查Excel数据中的错误
    
    Args:
        df: 待检查的DataFrame
        checks: 检查配置字典，如：
            {
                'check_empty_rows': True,  # 检查空行
                'check_null_values': True,  # 检查空值
                'check_duplicates': ['col1', 'col2'],  # 检查重复（指定列）
                'check_data_types': {'col1': 'int', 'col2': 'float'},  # 检查数据类型
                'check_value_range': {'col1': (0, 100)},  # 检查值范围
            }
        logger: 日志记录器，如果为None则使用默认logger
            
    Returns:
        (是否通过检查, 错误列表)
    """
    if logger is None:
        logger = _get_default_logger()
    
    if checks is None:
        checks = {
            'check_empty_rows': True,
            'check_null_values': True,
        }
    
    errors = []
    
    # 1. 检查空行
    if checks.get('check_empty_rows', False):
        empty_rows = df[df.isnull().all(axis=1)]
        if not empty_rows.empty:
            error = {
                'type': 'empty_rows',
                'count': len(empty_rows),
                'indices': empty_rows.index.tolist(),
                'message': f"发现 {len(empty_rows)} 个空行"
            }
            errors.append(error)
            logger.warning(error['message'])
    
    # 2. 检查空值
    if checks.get('check_null_values', False):
        null_counts = df.isnull().sum()
        null_cols = null_counts[null_counts > 0]
        if not null_cols.empty:
            error = {
                'type': 'null_values',
                'columns': null_cols.to_dict(),
                'message': f"发现空值: {null_cols.to_dict()}"
            }
            errors.append(error)
            logger.warning(error['message'])
    
    # 3. 检查重复行
    if 'check_duplicates' in checks:
        dup_cols = checks['check_duplicates']
        if dup_cols:
            duplicates = df[df.duplicated(subset=dup_cols, keep=False)]
            if not duplicates.empty:
                error = {
                    'type': 'duplicate_rows',
                    'count': len(duplicates),
                    'columns': dup_cols,
                    'message': f"发现 {len(duplicates)} 个重复行（基于列 {dup_cols}）"
                }
                errors.append(error)
                logger.warning(error['message'])
    
    # 4. 检查数据类型
    if 'check_data_types' in checks:
        type_checks = checks['check_data_types']
        for col, expected_type in type_checks.items():
            if col in df.columns:
                try:
                    if expected_type == 'int':
                        df[col].astype(int)
                    elif expected_type == 'float':
                        df[col].astype(float)
                    elif expected_type == 'datetime':
                        pd.to_datetime(df[col])
                except (ValueError, TypeError) as e:
                    error = {
                        'type': 'data_type_error',
                        'column': col,
                        'expected_type': expected_type,
                        'message': f"列 '{col}' 数据类型错误，期望 {expected_type}"
                    }
                    errors.append(error)
                    logger.warning(error['message'])
    
    # 5. 检查值范围
    if 'check_value_range' in checks:
        range_checks = checks['check_value_range']
        for col, (min_val, max_val) in range_checks.items():
            if col in df.columns:
                out_of_range = df[(df[col] < min_val) | (df[col] > max_val)]
                if not out_of_range.empty:
                    error = {
                        'type': 'value_out_of_range',
                        'column': col,
                        'range': (min_val, max_val),
                        'count': len(out_of_range),
                        'message': f"列 '{col}' 有 {len(out_of_range)} 个值超出范围 [{min_val}, {max_val}]"
                    }
                    errors.append(error)
                    logger.warning(error['message'])
    
    # 6. 检查异常值（使用IQR方法）
    if checks.get('check_outliers', False):
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = df[(df[col] < Q1 - 1.5 * IQR) | (df[col] > Q3 + 1.5 * IQR)]
            if not outliers.empty:
                error = {
                    'type': 'outliers',
                    'column': col,
                    'count': len(outliers),
                    'message': f"列 '{col}' 发现 {len(outliers)} 个异常值"
                }
                errors.append(error)
                logger.info(error['message'])
    
    passed = len(errors) == 0
    
    if passed:
        logger.info("Excel错误检查通过")
    else:
        logger.warning(f"Excel错误检查发现 {len(errors)} 个问题")
    
    return passed, errors
